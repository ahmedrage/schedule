#!python3
import requests, bs4, sys
import datetime
import time
from requests_futures.sessions import FuturesSession
from openpyxl import load_workbook, Workbook

retry=False
retries=0

#functions

#exports lecture to spreadsheet
def export(lectureItem):
    print(lectureItem)

    # Check if the file exists and if not create it

    try:
        workbook = load_workbook(filename="lectures.xlsx")
    except:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Course"
        sheet["B1"] = "Location"
        sheet["C1"] = "Time"
        sheet["D1"] = "Day"
        sheet["E1"] = "Type"
        workbook.save(filename="lectures.xlsx")
        workbook = load_workbook(filename="lectures.xlsx")
    sheet = workbook.active
    ind = sheet.max_row + 1
    ind = str(ind)
    sheet["A" + ind] = lectureItem[0]
    sheet["B" + ind] = lectureItem[1]
    sheet["C" + ind] = lectureItem[2]
    sheet["D" + ind] = lectureItem[3]
    sheet["E" + ind] = lectureItem[4]

    workbook.save(filename="lectures.xlsx")

#print lecture
#prints a lecture and its details e.g name, time, place
def printLec(lectureItem):
        lec = '('
        for i in lectureItem:
            lec += "'"
            lec += i
            lec += "', "
        lec += ')'
        print(lec)

#checks if a row in a table contains lectures
def isLecture(lec):
    lectureList = lec.findAll('td')
    if len(lectureList) > 3:
        return True
    else:
        return False

# def get_response_with_retry(target, future_session, sleep=5, retry=False, retries=0):
#     try:
#         request_to_send = None
#         if retry:
#             response = requests.get(target, timeout=60)
#         else:
#             request_to_send = future_session.get(url)
#             response = request_to_send.result()
#         response.raise_for_status()
#         print("response got")
#         return response
#     except Exception as ex:
#         if retries > 5:
#             raise Exception("Error: {0} - Giving up".format(ex))
#         print("Error: {0} - Retrying in {1} seconds. {2} retries".format(ex, sleep, retries))
#         time.sleep(sleep)
#         return get_response_with_retry(url, future_session, sleep, retry=True, retries=retries+1)
    
def get_response_with_retry(target, future_session, sleep=5):
    retry = False
    retries = 0
    for i in range(5):
        try:
            requset_to_send = None
            if retry:
                response = requests.get(target, timeout=60)
            else:
                request_to_send = future_session.get(target)
                response = request_to_send.result()
            response.raise_for_status()
            print("response got")
            return response
        except Exception as ex:
            print("Error: {0} - Retrying in {1} seconds. {2} retries".format(ex, sleep, retries))
            time.sleep(sleep)
            retry = True
            retries += 1
            continue

    raise Exception("Error: {0} - Giving up".format(ex))

#takes a lectures from a course page and stores them in the spreadsheet
def getLectures(res):

    #requesting page

    courseLectures =[]
    #Turn it into a BeautifulSoup
    courseSoup = bs4.BeautifulSoup(res.text, "html.parser")

    #find the table with lectures in it
    elems = courseSoup.select('title')
    courseTitle = elems[0].text
    elems = courseSoup.select('#hidedata04_1')
    if len(elems) == 0:
        print(courseTitle + ' ERROR: NO LECTURES FOUND')
        return False
    table = elems[0].contents
    table = table[1]

    #find lectures
    lectures = table.findAll('tr')
    if len(lectures) == 0 or lectures[0].getText() != u'Enrolment Class: Lecture':
        print(courseTitle + ' ERROR: NO LECTURES FOUND')
    else:
        lecture = lectures[2]
        lectureList = lecture.findAll('td')
        lectureLength = len(lectureList)

        #lecture item = (name ,location, time, day)

        #puts first lecture into list
        lectureItem = (courseTitle, lectureList[lectureLength-1].getText(), lectureList[lectureLength-2].getText(), lectureList[lectureLength-3].getText(), lectureList[lectureLength-4].getText())
        printLec(lectureItem)
        courseLectures.append(lectureItem)

        #checks if there are other lectures in page and adds them to list
        i = 3
        while i < len(lectures) and isLecture(lectures[i]):
            print(i)

            lecture = lectures[i]
            lectureList = lecture.findAll('td')
            lectureLength = len(lectureList)

            #lecture item = (name ,location, time, day)

            lectureItem = (courseTitle, lectureList[lectureLength-1].getText(), lectureList[lectureLength-2].getText(), lectureList[lectureLength-3].getText(),lectureList[lectureLength-4].getText())
            printLec(lectureItem)
            courseLectures.append(lectureItem)
            i = i + 1
    print("done")
    #export to spreadsheet

    for Lec in courseLectures:
        export(Lec)
    return True

#goes through the page of a subject and opens all the courses and then adds the lectures in them to the spread sheet.
def getCourses(res):
    #Turn it into a BeautifulSoup
    courseSoup = bs4.BeautifulSoup(res.text, "html.parser")
    links = []
    elems = courseSoup.findAll("table")
    table = elems[4]
    linkObj = table.findAll('a')
    switch = True

    for link in linkObj:
        if switch:
            links.append("https://access.adelaide.edu.au/courses/" + link['href'])
        switch = not switch

    future_session = FuturesSession()
    future_requests = []
    future_responses = []

    for url in links:
        try:
            res = get_response_with_retry(url, future_session)
        except Exception as ex:
            print("Error: {0}".format(ex))
            continue
        future_responses.append(res)
    
    # for req in future_requests:
    #     res = req.result()
    #     future_responses.append(res)

    for res in future_responses:
        getLectures(res)

#this part is the code run

#sends request form main page
# Get current year as string
current_year = str(datetime.datetime.now().year)
print("Searching for terms in year {0}".format(current_year))
main_url = "https://access.adelaide.edu.au/courses/search.asp?year={0}".format(current_year)
res = requests.get(main_url, timeout=10)
res.raise_for_status()

codes = []
terms = {}

#turns into soup
courseSoup = bs4.BeautifulSoup(res.text, "html.parser")

#finds all the different subject codes and puts them in a list
select = courseSoup.findAll("select")
optionObject = select[0].findAll("option")

for option in optionObject:
    if option['value'] != "":
        codes.append(option['value'])

select = courseSoup.findAll("select")
optionObject = select[1].findAll("option")

for option in optionObject:
    if option['value'] != "":
        terms[option.getText()] = option['value']
print("Select a term:")

for index, term in enumerate(terms.keys()):
    print("{0}: {1}".format(index, term))

selected_index = input()

selected_term = terms[list(terms.keys())[int(selected_index)]]
print(selected_term)

urls = []
async_list = []
#goes through that list and opens each of the subjects to extract lectures
for code in codes:
    urls.append("https://access.adelaide.edu.au/courses/search.asp?year={0}&m=r&title=&subject={1}&catalogue=&action=Search&term={2}".format(current_year, code, selected_term))

session = FuturesSession()
future_requests = []
future_responses = []

for url in urls:
    res = None
    try:
        res = get_response_with_retry(url, session)
    except Exception as ex:
        print("Error: {0}".format(ex))
        continue
    future_responses.append(res)
# for req in future_requests:
#     res = req.result()
#     future_responses.append(res)
#     print ("res got")

for res in future_responses:
    getCourses(res)


